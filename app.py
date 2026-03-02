from flask import Flask, request, send_file, jsonify
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# optional parsers
try:
    from pdfminer.high_level import extract_text as extract_text_from_pdf
except Exception:
    extract_text_from_pdf = None

try:
    import docx
except Exception:
    docx = None

app = Flask(__name__, static_folder='static', static_url_path='')


def create_class_sheet(wb: Workbook, cls: dict):
    name = cls.get('name', 'Class')[:31]
    raw_assigns = cls.get('assignments', [])

    # expand assignments according to count
    assignments = []
    for a in raw_assigns:
        cnt = a.get('count', 1) or 1
        w = a.get('weight', 0)
        # store weights as fractions (e.g. 2% -> 0.02) so formulas can multiply directly
        per = (w / 100.0) / cnt if cnt else 0
        for ii in range(cnt):
            aname = a.get('name', '')
            if cnt > 1:
                aname = f"{aname} {ii+1}"
            assignments.append({'name': aname, 'weight': per})

    ws = wb.create_sheet(title=name)

    # Header row (assignment names) + TOTAL column
    headers = [a['name'] for a in assignments] + ['TOTAL']
    ws.append([''] + headers)

    # Weight row (percent values stored as fraction so formulas can multiply directly)
    weight_row = 2
    ws.cell(row=weight_row, column=1, value='WEIGHT')
    for i, a in enumerate(assignments, start=2):
        cell = ws.cell(row=weight_row, column=i, value=a.get('weight', 0))
        cell.number_format = '0%'
    total_col = len(assignments) + 2
    first_col_letter = get_column_letter(2)
    last_col_letter = get_column_letter(len(assignments) + 1)
    total_cell = ws.cell(row=weight_row, column=total_col, value=f"=SUM({first_col_letter}{weight_row}:{last_col_letter}{weight_row})")
    total_cell.number_format = '0%'

    # Score row (user input, initialized to 0)
    score_row = 3
    ws.cell(row=score_row, column=1, value='SCORE')
    for i in range(2, len(assignments) + 2):
        ws.cell(row=score_row, column=i, value=0)
    # no SUM total for SCORE row per user request; leave blank
    ws.cell(row=score_row, column=total_col, value='')

    # Grade row: per-assignment contribution = SCORE * WEIGHT
    grade_row = 4
    ws.cell(row=grade_row, column=1, value='GRADE')
    for col in range(2, len(assignments) + 2):
        score_cell = ws.cell(row=score_row, column=col).coordinate
        weight_cell = ws.cell(row=weight_row, column=col).coordinate
        ws.cell(row=grade_row, column=col, value=f"={score_cell}*{weight_cell}")
    ws.cell(row=grade_row, column=total_col, value=f"=SUM({get_column_letter(2)}{grade_row}:{get_column_letter(len(assignments)+1)}{grade_row})")

    # Average score row (user input per assignment, initialized to 0)
    avgscore_row = 6
    ws.cell(row=avgscore_row, column=1, value='AVGSCORE')
    for i in range(2, len(assignments) + 2):
        ws.cell(row=avgscore_row, column=i, value=0)
    # no SUM total for AVGSCORE row per user request; leave blank
    ws.cell(row=avgscore_row, column=total_col, value='')

    # Average grade row: per-assignment contribution = AVGSCORE * WEIGHT
    avggrade_row = 7
    ws.cell(row=avggrade_row, column=1, value='AVGGRADE')
    for col in range(2, len(assignments) + 2):
        avgscore_cell = ws.cell(row=avgscore_row, column=col).coordinate
        weight_cell = ws.cell(row=weight_row, column=col).coordinate
        ws.cell(row=avggrade_row, column=col, value=f"={avgscore_cell}*{weight_cell}")
    ws.cell(row=avggrade_row, column=total_col, value=f"=SUM({get_column_letter(2)}{avggrade_row}:{get_column_letter(len(assignments)+1)}{avggrade_row})")

    # Note: per-user request, no overall TOTAL GRADE row is created.


def parse_weights_from_text(text: str):
    import re
    keywords = ['homework', 'midterm', 'final', 'assignment', 'quiz', 'exam', 'project', 'essay', 'presentation', 'lab']

    seen = set()
    out = []
    cumulative = 0

    # analyze line-by-line for any weight percentages, but only accept
    # candidates that contain one of the known assessment keywords. Stop
    # indexing once cumulative weight reaches or exceeds 100%.
    for line in text.splitlines():
        if '%' not in line:
            continue
        m = re.search(r"(\d{1,3})\s*%", line)
        if not m:
            continue
        try:
            weight = int(m.group(1))
        except ValueError:
            continue

        before = line[:m.start()].strip(' -:\t()[]')
        after = line[m.end():].strip(' -:\t()[]')

        candidate = None
        # prefer the token before the percent if it contains a keyword
        if before and any(k in before.lower() for k in keywords):
            candidate = before
        elif after and any(k in after.lower() for k in keywords):
            candidate = after
        else:
            # if neither side contains a keyword, check the whole line for keywords
            for k in keywords:
                if re.search(r"\b" + re.escape(k) + r"\b", line, re.I):
                    candidate = k.capitalize()
                    break

        if not candidate:
            continue

        name = candidate.strip()
        if name in seen:
            continue

        # append and update cumulative total; stop when reaching >= 100
        out.append({'name': name, 'weight': weight})
        seen.add(name)
        cumulative += weight
        if cumulative >= 100:
            break

    return out


@app.route('/')
def index():
    return app.send_static_file('index.html')


@app.route('/generate', methods=['POST'])
def generate():
    data = request.get_json(silent=True)
    if not data or 'classes' not in data:
        return jsonify({'error': 'Provide JSON with "classes" key.'}), 400

    classes = data.get('classes')
    if not isinstance(classes, list) or len(classes) == 0:
        return jsonify({'error': 'Add at least one class before generating Excel.'}), 400

    wb = Workbook()
    # remove the default sheet if present
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    for cls in classes:
        create_class_sheet(wb, cls)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(buf, as_attachment=True, download_name='gradebook.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/extract', methods=['POST'])
def extract():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file part'}), 400

        f = request.files['file']
        filename = f.filename or ''
        data = None

        # read raw bytes once so we can re-use them
        raw = f.read()

        # Try DOCX (only when extension is .docx and library available)
        if filename.lower().endswith('.docx') and docx is not None:
            try:
                from io import BytesIO
                doc = docx.Document(BytesIO(raw))
                data = '\n'.join(p.text for p in doc.paragraphs)
            except Exception:
                data = None

        # Try PDF (only when extension .pdf and extractor available)
        if data is None and filename.lower().endswith('.pdf') and extract_text_from_pdf is not None:
            try:
                from io import BytesIO
                data = extract_text_from_pdf(BytesIO(raw))
            except Exception as e:
                return jsonify({'error': f'PDF extraction failed: {str(e)}'}), 500

        # Fallback to plain text decoding
        if data is None:
            # try several encodings; prefer utf-8 but fall back to utf-16 if nulls present
            data = ''
            for enc in ('utf-8', 'utf-16', 'latin-1'):
                try:
                    candidate = raw.decode(enc)
                    if enc == 'utf-8' and '\x00' in candidate:
                        raise UnicodeDecodeError(enc, raw, 0, 1, 'nulls')
                    data = candidate
                    break
                except Exception:
                    continue
            if data == '':
                try:
                    data = raw.decode('utf-8', errors='ignore')
                except Exception:
                    data = ''

        weights = parse_weights_from_text(data or '')
        # return assignments plus the extracted text for debugging/manual fallback
        return jsonify({'assignments': weights, 'text': data})
    except Exception as e:
        return jsonify({'error': f'Unexpected server error: {str(e)}'}), 500


if __name__ == '__main__':
    app.run(debug=True, port=5000)
